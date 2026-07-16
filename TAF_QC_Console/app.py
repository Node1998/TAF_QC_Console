"""
app.py - TAF QC Console web service (Flask).

Deploy target: Render web service (gunicorn wsgi:app). Also runs locally:
    pip install -r requirements.txt
    python app.py

Storage: SQLite. Path configurable with TAF_DB_PATH (point it at a Render
persistent disk, e.g. /var/data/taf_validation.db; otherwise the DB is
ephemeral and resets on each deploy).
Exports: /api/export (CSV) and /api/export/xlsx (Excel, two sheets:
Validations + per-finding Findings rows).
"""
import os
import re
import csv
import json
import math
import logging
import functools
from io import StringIO, BytesIO
from datetime import datetime, timezone

import sqlite3
import requests
from flask import Flask, request, jsonify, Response, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from qc_engine import qc_taf

# =========================================================================== #
#  Configuration & Constants
# =========================================================================== #

APP_DIR  = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.environ.get("TAF_DB_PATH", os.path.join(APP_DIR, "taf_validation.db"))
AWC_TAF  = "https://aviationweather.gov/api/data/taf"
AWC_INFO = "https://aviationweather.gov/api/data/stationinfo"

# Upload constraints
MAX_UPLOAD_BYTES = 262_144   # 256 KB
MAX_UPLOAD_TAFS  = 50

# Regional consensus defaults
MAX_NEIGHBORS = 4
NEIGHBOR_BBOX_RADIUS = 1.5   # degrees
EARTH_RADIUS_NM = 3440.065   # nautical miles
DIRECTION_ANGLE_THRESHOLD = 180  # degrees

# Consensus scoring weights
SCORE_MAX = 100.0
DIR_DIFF_PENALTY = 40  # points per 180°
SPD_DIFF_PENALTY = 25  # max penalty for speed
VIS_DIFF_PENALTY = 35  # max penalty for visibility
NO_DATA_PENALTY = 15   # penalty when data missing

# HTTP timeouts
AWC_INFO_TIMEOUT = 6     # seconds
AWC_TAF_TIMEOUT = 10     # seconds
AWC_STATUS_TIMEOUT = 10  # seconds

# Precompiled regex patterns
ICAO_PATTERN = re.compile(r"[A-Z0-9]{4}")
WIND_PATTERN = re.compile(r"\b(VRB|\d{3})(\d{2,3})(?:G\d{2,3})?KT\b")
VISIBILITY_PATTERN = re.compile(r"\b(\d{4})\b(?=\s|$)")
TAF_SPLIT_PATTERN = re.compile(r"(?=\bTAF\s)", re.I)
TAF_BLOCK_PATTERN = re.compile(r"\n\s*\n")

app = Flask(__name__)

# Configure logging
logger = logging.getLogger(__name__)
handler = logging.StreamHandler()
handler.setFormatter(logging.Formatter(
    "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
))
logger.addHandler(handler)
logger.setLevel(logging.INFO)


# =========================================================================== #
#  Utility Functions
# =========================================================================== #

def _validate_icao(icao_str):
    """Validate and normalize an ICAO code.
    Returns normalized code or None if invalid."""
    icao = (icao_str or "").strip().upper()
    if not ICAO_PATTERN.fullmatch(icao):
        return None
    return icao


def _escape_csv_formula(val):
    """Prevent CSV injection by escaping formula-like strings.
    Excel treats =, +, @, - as formula prefixes when at cell start."""
    if isinstance(val, str) and val and val[0] in '=+-@':
        return f"'{val}"
    return val


# =========================================================================== #
#  Storage layer (SQLite, parameterized queries only)
# =========================================================================== #

def _conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    """Initialize database schema."""
    conn = _conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS _migrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            applied_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS taf_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            icao TEXT, issue_time TEXT, validity TEXT, wind TEXT, visibility TEXT,
            clouds TEXT, altimeter TEXT, qc_status TEXT, qc_score INTEGER,
            findings TEXT, source TEXT, raw_text TEXT, logged_at TEXT
        )
    """)
    conn.commit()
    conn.close()
    logger.info("Database initialized")


def _has_migration_run(conn, migration_name):
    """Check if a migration has already been applied."""
    result = conn.execute(
        "SELECT id FROM _migrations WHERE name = ?",
        (migration_name,)
    ).fetchone()
    return result is not None


def _record_migration(conn, migration_name):
    """Record a completed migration."""
    conn.execute(
        "INSERT INTO _migrations (name, applied_at) VALUES (?, ?)",
        (migration_name, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ"))
    )


def upgrade_db():
    """Apply idempotent migrations."""
    conn = _conn()

    # Migration 1: Create taf_findings table
    if not _has_migration_run(conn, "001_create_taf_findings"):
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS taf_findings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    log_id INTEGER NOT NULL REFERENCES taf_logs(id) ON DELETE CASCADE,
                    severity TEXT NOT NULL,
                    ref TEXT,
                    message TEXT
                )
            """)
            _record_migration(conn, "001_create_taf_findings")
            logger.info("Migration 001_create_taf_findings applied")
        except Exception as e:
            logger.error(f"Migration 001 failed: {e}")

    # Migration 2: Add weather column to taf_logs
    if not _has_migration_run(conn, "002_add_weather_column"):
        try:
            cols = {r["name"] for r in conn.execute("PRAGMA table_info(taf_logs)")}
            if "weather" not in cols:
                conn.execute("ALTER TABLE taf_logs ADD COLUMN weather TEXT")
                logger.info("Added weather column to taf_logs")
            _record_migration(conn, "002_add_weather_column")
        except Exception as e:
            logger.error(f"Migration 002 failed: {e}")

    # Migration 3: Add penalty and token columns to taf_findings
    if not _has_migration_run(conn, "003_add_finding_columns"):
        try:
            fcols = {r["name"] for r in conn.execute("PRAGMA table_info(taf_findings)")}
            if "penalty" not in fcols:
                conn.execute("ALTER TABLE taf_findings ADD COLUMN penalty INTEGER DEFAULT 0")
                logger.info("Added penalty column to taf_findings")
            if "token" not in fcols:
                conn.execute("ALTER TABLE taf_findings ADD COLUMN token TEXT")
                logger.info("Added token column to taf_findings")
            _record_migration(conn, "003_add_finding_columns")
        except Exception as e:
            logger.error(f"Migration 003 failed: {e}")

    conn.commit()
    conn.close()
    logger.info("Database migrations complete")


def insert_log(g, qc, source, raw):
    """Insert a validation log and its findings into the database."""
    conn = _conn()
    try:
        cur = conn.execute("""
            INSERT INTO taf_logs (icao, issue_time, validity, wind, visibility, clouds,
                weather, altimeter, qc_status, qc_score, findings, source, raw_text, logged_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (g.get("icao"), g.get("issue"), g.get("valid"), g.get("wind"),
              g.get("visibility"), g.get("clouds"), g.get("weather"), g.get("altimeter"),
              qc["status"], qc["score"], json.dumps(qc["findings"]), source, raw,
              datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")))
        rid = cur.lastrowid
        
        conn.executemany(
            "INSERT INTO taf_findings (log_id, severity, ref, message, penalty, token) "
            "VALUES (?,?,?,?,?,?)",
            [(rid, f["severity"], f["ref"], f["message"],
              f.get("penalty", 0), f.get("token")) for f in qc["findings"]])
        conn.commit()
        logger.info(f"Logged TAF validation: id={rid}, icao={g.get('icao')}, source={source}")
        return rid
    except Exception as e:
        logger.error(f"Failed to insert log: {e}")
        raise
    finally:
        conn.close()


def fetch_logs(limit=200):
    """Fetch the most recent validation logs."""
    conn = _conn()
    try:
        rows = conn.execute("SELECT * FROM taf_logs ORDER BY id DESC LIMIT ?",
                            (limit,)).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"Failed to fetch logs: {e}")
        return []
    finally:
        conn.close()


LOG_COLS = ("id", "icao", "issue_time", "validity", "wind", "visibility",
            "weather", "clouds", "altimeter", "qc_status", "qc_score",
            "source", "raw_text", "logged_at")
FIND_COLS = ("log_id", "icao", "qc_status", "severity", "ref", "penalty",
             "token", "message")


def _export_rows():
    """Fetch all validation logs and findings for export."""
    conn = _conn()
    try:
        # Use quoted column names to avoid injection risks
        log_cols_quoted = ", ".join(f'"{c}"' for c in LOG_COLS)
        logs = conn.execute(
            f"SELECT {log_cols_quoted} FROM taf_logs ORDER BY id ASC").fetchall()
        finds = conn.execute("""
            SELECT f.log_id, l.icao, l.qc_status, f.severity, f.ref,
                   f.penalty, f.token, f.message
            FROM taf_findings f JOIN taf_logs l ON l.id = f.log_id
            ORDER BY f.log_id ASC, f.id ASC""").fetchall()
        return logs, finds
    except Exception as e:
        logger.error(f"Failed to export rows: {e}")
        return [], []
    finally:
        conn.close()


def export_csv():
    """Export validation logs as CSV with formula injection protection."""
    logs, _ = _export_rows()
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(LOG_COLS)
    for r in logs:
        # Escape formula-like strings to prevent CSV injection
        row = [_escape_csv_formula(r[c]) for c in LOG_COLS]
        w.writerow(row)
    logger.info(f"Exported {len(logs)} logs to CSV")
    return buf.getvalue()


def export_xlsx():
    """Export validation logs and findings as Excel workbook."""
    logs, finds = _export_rows()
    wb = Workbook()

    head_fill = PatternFill("solid", fgColor="111C2E")
    head_font = Font(bold=True, color="FFFFFF")
    status_fill = {"PASS": "3DDC84", "REVIEW": "F5A623", "FAIL": "FF5D5D"}

    # Validations sheet
    ws = wb.active
    ws.title = "Validations"
    ws.append([c.upper() for c in LOG_COLS])
    for cell in ws[1]:
        cell.fill, cell.font = head_fill, head_font
    st_col = LOG_COLS.index("qc_status") + 1
    for r in logs:
        # Escape formula-like strings
        row = [_escape_csv_formula(r[c]) for c in LOG_COLS]
        ws.append(row)
        fill = status_fill.get(r["qc_status"])
        if fill:
            ws.cell(row=ws.max_row, column=st_col).fill = PatternFill("solid", fgColor=fill)
    for col, width in zip("ABCDEFGHIJKLMN", (6, 8, 10, 11, 14, 11, 14, 18, 10, 10, 9, 40, 22)):
        ws.column_dimensions[col].width = width

    # Findings sheet
    ws2 = wb.create_sheet("Findings")
    ws2.append([c.upper() for c in FIND_COLS])
    for cell in ws2[1]:
        cell.fill, cell.font = head_fill, head_font
    for r in finds:
        row = [_escape_csv_formula(r[c]) for c in FIND_COLS]
        ws2.append(row)
    for col, width in zip("ABCDEFGH", (8, 8, 10, 10, 18, 9, 14, 80)):
        ws2.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"Exported {len(logs)} logs and {len(finds)} findings to Excel")
    return buf


# =========================================================================== #
#  Data ingestion + optional regional consensus
# =========================================================================== #

def fetch_taf(icao):
    """Fetch a TAF from Aviation Weather Center."""
    icao_code = _validate_icao(icao)
    if not icao_code:
        logger.warning(f"Invalid ICAO format: {icao}")
        return None
    try:
        r = requests.get(AWC_TAF, params={"ids": icao_code, "format": "raw"}, 
                        timeout=AWC_TAF_TIMEOUT)
        r.raise_for_status()
        result = r.text.strip() or None
        if result:
            logger.info(f"Fetched TAF for {icao_code}")
        else:
            logger.warning(f"No TAF returned for {icao_code}")
        return result
    except requests.exceptions.Timeout:
        logger.error(f"Timeout fetching TAF for {icao_code}")
        return None
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching TAF for {icao_code}: {e}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error fetching TAF for {icao_code}: {e}")
        return None


def _haversine_nm(lat1, lon1, lat2, lon2):
    """Calculate great-circle distance between two points in nautical miles."""
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1))
         * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2)
    return EARTH_RADIUS_NM * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


@functools.lru_cache(maxsize=128)
def _parse_wind_visibility(raw):
    """Parse wind and visibility from a TAF text.
    Cached to avoid recompiling regex on every call.
    Returns (direction, speed, visibility) tuple."""
    wm = WIND_PATTERN.search(raw or "")
    vm = VISIBILITY_PATTERN.search((raw or "").split("KT", 1)[-1])
    d = None if (not wm or wm.group(1) == "VRB") else int(wm.group(1))
    s = int(wm.group(2)) if wm else None
    v = int(vm.group(1)) if vm else 9999
    return d, s, min(v, 9999)


def regional_consensus(target_icao, target_raw):
    """Compare the target TAF's wind/visibility to its nearest neighbors.
    Returns (neighbors, leaderboard). Network failures degrade to ([], [])."""
    try:
        info_resp = requests.get(AWC_INFO, 
                                params={"ids": target_icao, "format": "json"},
                                timeout=AWC_INFO_TIMEOUT)
        info_resp.raise_for_status()
        info = info_resp.json()
        
        if not info:
            logger.warning(f"No station info for {target_icao}")
            return [], []
        
        tlat, tlon = float(info[0].get("lat", 0)), float(info[0].get("lon", 0))
        if tlat == 0 and tlon == 0:
            logger.warning(f"Invalid coordinates for {target_icao}")
            return [], []
        
        bbox = f"{tlat-NEIGHBOR_BBOX_RADIUS},{tlon-NEIGHBOR_BBOX_RADIUS}," \
               f"{tlat+NEIGHBOR_BBOX_RADIUS},{tlon+NEIGHBOR_BBOX_RADIUS}"
        
        tafs_resp = requests.get(AWC_TAF, params={"bbox": bbox, "format": "json"},
                                timeout=AWC_TAF_TIMEOUT)
        tafs_resp.raise_for_status()
        tafs = tafs_resp.json()
    except requests.exceptions.Timeout:
        logger.error(f"Timeout fetching consensus data for {target_icao}")
        return [], []
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching consensus data for {target_icao}: {e}")
        return [], []
    except Exception as e:
        logger.error(f"Unexpected error in regional_consensus for {target_icao}: {e}")
        return [], []

    neighbors, seen = [], set()
    for t in tafs:
        nic = t.get("icaoId")
        if not nic or nic == target_icao or nic in seen:
            continue
        nlat, nlon = float(t.get("lat", 0)), float(t.get("lon", 0))
        if not (nlat and nlon):
            continue
        seen.add(nic)
        neighbors.append({"icao": nic,
                          "dist": round(_haversine_nm(tlat, tlon, nlat, nlon), 1),
                          "raw": t.get("rawTAF", "")})
    
    neighbors.sort(key=lambda x: x["dist"])
    neighbors = neighbors[:MAX_NEIGHBORS]
    logger.info(f"Found {len(neighbors)} neighbors for {target_icao}")

    rows = [{"icao": target_icao, **dict(zip(("dir", "spd", "vis"), 
            _parse_wind_visibility(target_raw)))}]
    for n in neighbors:
        rows.append({"icao": n["icao"], **dict(zip(("dir", "spd", "vis"), 
                    _parse_wind_visibility(n["raw"])))})

    dirs = [r["dir"] for r in rows if r["dir"] is not None]
    spds = [r["spd"] for r in rows if r["spd"] is not None]
    viss = [r["vis"] for r in rows if r["vis"] is not None]
    avg_spd = sum(spds) / len(spds) if spds else 0
    avg_vis = sum(viss) / len(viss) if viss else 9999
    
    if dirs:
        s = sum(math.sin(math.radians(d)) for d in dirs)
        c = sum(math.cos(math.radians(d)) for d in dirs)
        avg_dir = (math.degrees(math.atan2(s, c)) + 360) % 360
    else:
        avg_dir = 0

    board = []
    for r in rows:
        sc = SCORE_MAX
        if r["dir"] is not None and dirs:
            diff = abs(r["dir"] - avg_dir)
            diff = DIRECTION_ANGLE_THRESHOLD - diff if diff > 180 else diff
            sc -= diff / 180.0 * DIR_DIFF_PENALTY
        else:
            sc -= NO_DATA_PENALTY
        if r["spd"] is not None and spds:
            sc -= min(SPD_DIFF_PENALTY, abs(r["spd"] - avg_spd) * 2)
        if r["vis"] is not None and viss:
            sc -= min(VIS_DIFF_PENALTY, abs(r["vis"] - avg_vis) / 1000 * 5)
        board.append({"icao": r["icao"], "score": round(max(0, sc), 1)})
    
    board.sort(key=lambda x: x["score"], reverse=True)
    for i, b in enumerate(board):
        b["rank"] = i + 1
    
    return neighbors, board


# =========================================================================== #
#  Global JSON error handlers - the API must never return HTML error pages.
# =========================================================================== #

@app.errorhandler(Exception)
def _handle_exception(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        logger.warning(f"HTTP error {e.code}: {e.description}")
        return jsonify({"ok": False, "error": e.description}), e.code
    logger.exception("Unhandled exception")
    return jsonify({"ok": False, "error": f"Internal error: {type(e).__name__}"}), 500


# =========================================================================== #
#  Routes
# =========================================================================== #

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/process", methods=["POST"])
def api_process():
    try:
        p = request.get_json(force=True, silent=True) or {}
    except Exception as e:
        logger.warning(f"Invalid JSON in /api/process: {e}")
        return jsonify({"ok": False, "error": "Invalid JSON payload"}), 400
    
    mode = p.get("mode", "live")
    icao = (p.get("icao") or "").strip().upper()

    if mode == "live":
        raw = fetch_taf(icao)
        if not raw:
            return jsonify({"ok": False,
                            "error": f"No TAF returned for '{icao}'. "
                                     "Check the ICAO or your connection."})
        source = "LIVE"
    else:
        raw = (p.get("raw") or "").strip()
        if not raw:
            return jsonify({"ok": False, "error": "Paste a TAF to validate."})
        source = "MANUAL"

    qc = qc_taf(raw)
    g = qc["groups"]
    icao = icao or g.get("icao") or ""
    rid = insert_log(g, qc, source, raw)

    neighbors, board = ([], [])
    if p.get("consensus") and _validate_icao(icao):
        neighbors, board = regional_consensus(icao, raw)

    return jsonify({"ok": True, "id": rid, "source": source, "raw": raw,
                    "qc": qc, "neighbors": neighbors, "leaderboard": board})


def _split_tafs(blob):
    """Split an uploaded text blob into individual TAF messages.
    Blank-line separated blocks; a single block containing several 'TAF '
    headers is split before each header."""
    blocks = [b.strip() for b in TAF_BLOCK_PATTERN.split(blob) if b.strip()]
    if len(blocks) == 1 and blocks[0].upper().count("TAF ") > 1:
        parts = TAF_SPLIT_PATTERN.split(blocks[0])
        blocks = [p.strip() for p in parts if p.strip()]
    return blocks[:MAX_UPLOAD_TAFS]


@app.route("/api/upload", methods=["POST"])
def api_upload():
    f = request.files.get("file")
    if not f:
        logger.warning("Upload attempted without file")
        return jsonify({"ok": False, "error": "No file supplied."}), 400
    
    # Check content length before reading
    if request.content_length and request.content_length > MAX_UPLOAD_BYTES:
        logger.warning(f"File too large: {request.content_length} > {MAX_UPLOAD_BYTES}")
        return jsonify({"ok": False, "error": "File exceeds 256 KB limit."}), 413
    
    blob = f.read(MAX_UPLOAD_BYTES + 1)
    if len(blob) > MAX_UPLOAD_BYTES:
        logger.warning(f"File exceeds limit: {len(blob)} bytes")
        return jsonify({"ok": False, "error": "File exceeds 256 KB limit."}), 413
    
    try:
        textblob = blob.decode("utf-8", errors="replace")
    except Exception as e:
        logger.warning(f"File decode error: {e}")
        return jsonify({"ok": False, "error": "File is not readable text."}), 400
    
    tafs = _split_tafs(textblob)
    if not tafs:
        logger.warning("No TAF text found in uploaded file")
        return jsonify({"ok": False, "error": "No TAF text found in file."})
    
    results = []
    for raw in tafs:
        qc = qc_taf(raw)
        rid = insert_log(qc["groups"], qc, "UPLOAD", raw)
        results.append({"id": rid, "raw": raw, "qc": qc})
    
    logger.info(f"Processed {len(results)} TAFs from upload")
    return jsonify({"ok": True, "count": len(results), "results": results})


@app.route("/api/logs")
def api_logs():
    return jsonify({"ok": True, "logs": fetch_logs()})


@app.route("/api/export")
def api_export():
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return Response(export_csv(), mimetype="text/csv",
                    headers={"Content-Disposition":
                             f"attachment; filename=taf_logs_{stamp}.csv"})


@app.route("/api/export/xlsx")
def api_export_xlsx():
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return send_file(export_xlsx(), as_attachment=True,
                     download_name=f"taf_logs_{stamp}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    upgrade_db()
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", 5000)), debug=False)
